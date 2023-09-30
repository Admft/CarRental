import openpyxl
from datetime import datetime

class CarRentalSystem:
    def __init__(self, file_name):
        self.file_name = file_name
        self.headers = ['CarID', 'CarName', 'DailyRate', 'RentedBy', 'NumberOfDays', 'DateRented']
        self.init_excel()

    def init_excel(self):
        try:
            self.load_data()
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(self.headers)
            self.save_data()

    def load_data(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.ws = self.wb.active

    def save_data(self):
        self.wb.save(self.file_name)

    def add_car(self, car_name, daily_rate):
        self.load_data()
        car_id = str(len(list(self.ws.iter_rows()))).zfill(5)
        self.ws.append([car_id, car_name, daily_rate, None, None, None])
        self.save_data()

    def delete_car(self, car_id):
        self.load_data()
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == car_id:
                self.ws.delete_rows(row[0].row)
                self.save_data()
                return True
        return False

    def list_cars(self, available_only=False):
        self.load_data()
        cars = []
        print(f"{'Number':<10}{'CarID':<10}{'CarName':<20}{'DailyRate':<10}{'RentedBy':<20}{'NumberOfDays':<15}{'DateRented':<15}")
        print("-" * 90)
        for idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), 1):
            if not available_only or (available_only and row[3] is None):
                cars.append(row[0])
                print(f"{idx:<10}{row[0]:<10}{row[1]:<20}{row[2]:<10}{row[3] if row[3] else '':<20}{row[4] if row[4] else '':<15}{row[5] if row[5] else '':<15}")
        return cars

    def rent_car(self):
        available_cars = self.list_cars(available_only=True)
        if not available_cars:
            print("No cars available for rent!")
            return
        car_number = int(input("Enter the number of the car you want to rent: "))
        car_id = available_cars[car_number - 1]
        customer_name = input("Enter your name: ")
        days = int(input("Enter number of days: "))
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == car_id:
                row[3].value = customer_name
                row[4].value = days
                row[5].value = datetime.now().strftime("%Y-%m-%d")
                self.save_data()
                return True
        return False

    def return_car(self, car_id):
        self.load_data()
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == car_id and row[3].value is not None:
                row[3].value = None
                row[4].value = None
                row[5].value = None
                self.save_data()
                return True
        return False

if __name__ == '__main__':
    rental_system = CarRentalSystem("rental_data.xlsx")

    while True:
        print("\nOptions: \n1. List all cars \n2. Rent a car \n3. Return a car \n4. Add a car \n5. Delete a car \n6. Exit")
        choice = input("Enter your choice: ")

        if choice == "1":
            rental_system.list_cars()
        elif choice == "2":
            if rental_system.rent_car():
                print("Car rented successfully!")
            else:
                print("Error renting car.")
        elif choice == "3":
            car_id = input("Enter car ID to return: ")
            if rental_system.return_car(car_id):
                print("Car returned successfully!")
            else:
                print("Error returning car.")
        elif choice == "4":
            car_name = input("Enter car name: ")
            daily_rate = float(input("Enter daily rate: "))
            rental_system.add_car(car_name, daily_rate)
        elif choice == "5":
            car_id = input("Enter car ID to delete: ")
            if rental_system.delete_car(car_id):
                print("Car deleted successfully!")
            else:
                print("Error deleting car.")
        elif choice == "6":
            break
